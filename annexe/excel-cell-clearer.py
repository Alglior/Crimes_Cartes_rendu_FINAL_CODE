import os
from openpyxl import load_workbook

def remove_excel_cells(directory='per_sheets'):
    # Ensure the directory exists
    if not os.path.isdir(directory):
        print(f"The directory {directory} does not exist.")
        return

    # Get all .xlsx files in the directory
    excel_files = [f for f in os.listdir(directory) if f.endswith('.xlsx')]

    for file in excel_files:
        file_path = os.path.join(directory, file)
        
        try:
            # Load the workbook
            wb = load_workbook(file_path)
            sheet = wb.active

            # Clear cells from A1 to P8
            for row in range(1, 9):  # 1 to 8
                for col in range(1, 17):  # A to P
                    sheet.cell(row=row, column=col).value = None

            # Shift cells up
            for col in range(1, sheet.max_column + 1):
                column_cells = [cell.value for cell in sheet[chr(64 + col)][8:]]
                for row, value in enumerate(column_cells, start=1):
                    sheet.cell(row=row, column=col).value = value

            # Delete the now-empty rows at the bottom
            sheet.delete_rows(sheet.max_row - 7, 8)

            # Save the changes
            wb.save(file_path)
            print(f"Removed cells A1:P8 in '{file}' and shifted remaining cells up")

        except Exception as e:
            print(f"Error processing '{file}': {str(e)}")

if __name__ == "__main__":
    remove_excel_cells()