import openpyxl
from openpyxl import Workbook
import os

def separate_sheets(input_file):
    # Load the workbook
    wb = openpyxl.load_workbook(input_file)
    
    # Get the base filename without extension
    base_filename = os.path.splitext(input_file)[0]
    
    # Iterate through all sheets
    for sheet_name in wb.sheetnames:
        # Create a new workbook
        new_wb = Workbook()
        # Get the active sheet of the new workbook
        new_sheet = new_wb.active
        
        # Copy the sheet to the new workbook
        for row in wb[sheet_name].iter_rows(values_only=True):
            new_sheet.append(row)
        
        # Set the title of the new sheet
        new_sheet.title = sheet_name
        
        # Save the new workbook
        new_filename = f"{base_filename}_{sheet_name}.xlsx"
        new_wb.save(new_filename)
        print(f"Created: {new_filename}")

# Example usage
input_file = "crim_off_cat_spreadsheet.xlsx"
separate_sheets(input_file)